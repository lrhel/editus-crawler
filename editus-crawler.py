#!/usr/bin/env python3

import argparse
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from selenium import webdriver

def argparser():
    parser = argparse.ArgumentParser(prog = "Editus Crawler", description = "Crawl the result of a query on Editus.lu and put the result in an Excel Sheet")
    parser.add_argument("query", help="Query to crawl")
    parser.add_argument("output", help = "The output name")
    args = parser.parse_args()
    return args

def get_page(query, page):
    driver = webdriver.Firefox()
    driver.get("https://www.editus.lu/fr/recherche?q=" + query + "&p=" + str(page))
    return driver

def get_new_page(page, query, number):
    page.get("https://www.editus.lu/fr/recherche?q=" + query + "&p=" + str(number))
    return page

def get_total_page(page):    
    soup = BeautifulSoup(page.page_source, "lxml")
    for ul in soup.find_all('ul', {'class': 'pagination '}):
        return ul.find_all('li')[-2]
        

def main():
    wb = Workbook()
    ws = wb.active
    args = argparser()
    page = get_page(args.query, 1)
    total_page = int(get_total_page(page).find('a').get_text().strip(' '))
    for i in range(1, total_page + 1):
        page = get_new_page(page, args.query, i)
        soup = BeautifulSoup(page.page_source, "lxml")
        j = 0
        for main_info in soup.find_all('div', {'class': 'client-block'}):
            name = main_info.find('h2', {'class': 'name'}).find('a')
            address = main_info.find('h2', {'class': 'address'})
            phone_number = main_info.find('span', {'class': 'phone-number'})
            website = main_info.find('a', {'class': 'website button button-grey'})
            if phone_number:
                ws.cell(column=4, row=i+j*total_page+1, value=phone_number.get_text())
            if website:
                ws.cell(column=5, row=i+j*total_page+1, value=website.get('href'))
            ws.cell(column=1, row=i+j*total_page+1, value=name.get_text().strip(' '))
            ws.cell(column=2, row=i+j*total_page+1, value=name.get('href'))
            ws.cell(column=3, row=i+j*total_page+1, value=address.get_text())
            j = j + 1
    ws.cell(column=1, row=1, value="Name")
    ws.cell(column=2, row=1, value="Editus URL")
    ws.cell(column=3, row=1, value="Address")
    ws.cell(column=4, row=1, value="Phone Number")
    ws.cell(column=5, row=1, value="Website")
    wb.save(filename = args.output)

if __name__ == '__main__':
    main()
